__author__ = 'tzonliang.hsieh'

import time
import os
import openpyxl
import sys
import string
from openpyxl import load_workbook

class ExcelxlXrw:
    def __init__(self):
        print "Init class"

    def CheckFileVail(self,filepath):
        #filepath_str = "".join(filepath)
        #print filepath_str
        try:
            print os.path.abspath(str(filepath))
            wb2 = openpyxl.load_workbook(os.path.abspath(filepath))

        except Exception as inst:
            print "Open file and modify fail on :"+inst.strerror
            return False
        else:
            print "Can boot"
            return True
    def ModifyExcelData(self,filepath,sheet,col,row,data):
        #open file and modify
        filepath=filepath.strip('\n')
        #Change dir to xlsx file path and open file in abspath.
        os.path.dirname(filepath)
        try:
            wb2 = openpyxl.load_workbook(os.path.abspath(filepath))
            wa = wb2.active
            wa[self.xlsx_Table_Tran(col,row)].value = data
        # make sure excel file closed before save from code
            wb2.save(filepath)
            return True
        except Exception as inst:
            print "Open file and modify fail on :"+inst.strerror

    def xlsx_Table_Tran(self,col,row):
        Alpha=string.uppercase[:]
        try:
            return Alpha[col]+str(row)
        except Exception as inst:
            print "Table Trans fail"+inst.strerror
            return False


Times = time.time()
CheckDateStr = 'Check Date:' + time.strftime('%Y/%m/%d',time.localtime(Times))
OpenExcelObj = ExcelxlXrw()
print OpenExcelObj.__class__
if OpenExcelObj.CheckFileVail('Score card checklist for Self-Testing.xlsx1'):
    OpenExcelObj.ModifyExcelData('Score card checklist for Self-Testing.xlsx',sheet=0,col=3,row=2,data=CheckDateStr)
